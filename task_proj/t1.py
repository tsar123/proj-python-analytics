import csv
from _datetime import datetime
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from openpyxl.styles import Font
from openpyxl.styles.numbers import BUILTIN_FORMATS

dicKeys = {"name": "Название", "description": "Описание", "key_skills": "Навыки", "experience_id": "Опыт работы",
              "premium": "Премиум-вакансия", "employer_name": "Компания", "salary_to": "Оклад",
              "area_name": "Название региона",
              "published_at": "Дата публикации вакансии", "True": "Да", "False": "Нет", "FALSE": "FALSE",
              "TRUE": "TRUE",
              "value": "Идентификатор валюты оклада"}
reverse_dic_naming = {v: k for k, v in dicKeys.items()}
dicExperience = {"noExperience": "Нет опыта", "between1And3": "От 1 года до 3 лет",
                   "between3And6": "От 3 до 6 лет", "moreThan6": "Более 6 лет"}
workExperience = {"Нет опыта": 1, "От 1 года до 3 лет": 2, "От 3 до 6 лет": 3, "Более 6 лет": 4}
dicMoney = {"AZN": "Манаты", "BYR": "Белорусские рубли", "EUR": "Евро", "GEL": "Грузинский лари",
              "KGS": "Киргизский сом",
              "KZT": "Тенге", "RUR": "Рубли", "UAH": "Гривны", "USD": "Доллары", "UZS": "Узбекский сум"}
currencyToRub = {
    "AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
    "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}
dicHeadSecond = ['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']

class Vacancy:
    def __init__(self, name, salary, area_name, published_at):
        self.name = name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at


class Salary:
    def __init__(self, salary_from, salary_to, salary_currency):
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency
        self.salary_ru = int((float(self.salary_from) + float(self.salary_to)) / 2) * currencyToRub[
            self.salary_currency]

    def get_salary_ru(self):
        return self.salary_ru


class DataSet:
    def __init__(self, file_name):
        self.file_name = file_name
        self.vacancies_objects = DataSet.prepare(file_name)

    @staticmethod
    def csv_reader(filename):
        with open(filename, encoding="utf-8-sig") as f:
            data = [x for x in csv.reader(f)]
        try:
            clmns = data[0]
            lines = data[1:]
            return clmns, lines
        except FileNotFoundError:
            print("Пустой файл")
            exit()

    @staticmethod
    def prepare(filename):
        clmns, lines = DataSet.csv_reader(filename)
        filtred = [i for i in lines if len(i) == len(clmns) and '' not in i]
        vac = []
        for line in filtred:
            dct = {}
            for x in range(0, len(line)):
                if line[x].count('\n') > 0:
                    read = [DataSet.remove_tags(el) for el in line[x].split('\n')]
                else:
                    read = DataSet.remove_tags(line[x])
                dct[clmns[x]] = read

            vac.append(Vacancy(dct['name'], Salary(dct['salary_from'],
                                                   dct['salary_to'], dct['salary_currency']), dct['area_name'],
                               dct['published_at']))
        return vac

    @staticmethod
    def remove_tags(args):
        return " ".join(re.sub(r"\<[^>]*\>", "", args).split())


class InputParam:
    def __init__(self):
        self.params = InputParam.get_param()

    @staticmethod
    def get_param():
        file_name = "vacancies_with_skills.csv"
        vacancy = "devops"
        return file_name, vacancy

    @staticmethod
    def first_corr(dic):
        res = {}
        i = 0
        for key, value in dic.items():
            res[key] = value
            i += 1
            if i == 10:
                break
        return res

    @staticmethod
    def print(dic_vacancies, vac_name):
        years = set()
        for vacancy in dic_vacancies:
            years.add(int(datetime.strptime(vacancy.published_at, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y')))
        years = sorted(list(years))
        years = list(range(min(years), max(years) + 1))

        salary_years = {year: [] for year in years}
        vacs_years = {year: 0 for year in years}
        vac_salary_years = {year: [] for year in years}
        vac_count_years = {year: 0 for year in years}

        for vacancy in dic_vacancies:
            year = int(datetime.strptime(vacancy.published_at, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y'))
            salary_years[year].append(vacancy.salary.get_salary_ru())
            vacs_years[year] += 1
            if vac_name in vacancy.name:
                vac_salary_years[year].append(vacancy.salary.get_salary_ru())
                vac_count_years[year] += 1

        salary_years = {key: int(sum(value) / len(value)) if len(value) != 0 else 0
                        for key, value in salary_years.items()}
        vac_salary_years = {key: int(sum(value) / len(value)) if len(value) != 0 else 0
                            for key, value in vac_salary_years.items()}

        area_dic = {}
        for vacancy in dic_vacancies:
            if vacancy.area_name in area_dic:
                area_dic[vacancy.area_name].append(vacancy.salary.get_salary_ru())
            else:
                area_dic[vacancy.area_name] = [vacancy.salary.get_salary_ru()]

        area_salary = [x for x in area_dic.items() if len(x[1]) / len(dic_vacancies) > 0.01]
        sort_area_salary = sorted(area_salary, key=lambda item: sum(item[1]) / len(item[1]), reverse=True)
        res_sort_area_salary = {item[0]: int(sum(item[1]) / len(item[1])) for item in sort_area_salary}

        fract_vac_area = {
            key: round(len(value) / len(dic_vacancies), 4) if len(value) / len(dic_vacancies) > 0.01 else 0
            for key, value in area_dic.items()}
        fract_vac_area = {key: value for key, value in fract_vac_area.items() if value != 0}
        sort_fract_vac_area = sorted(fract_vac_area.items(), key=lambda item: item[1], reverse=True)
        res_sort_fract_vac_area = {k: v for k, v in sort_fract_vac_area}

        print('Динамика уровня зарплат по годам: {}'.format(salary_years))
        print('Динамика количества вакансий по годам: {}'.format(vacs_years))
        print('Динамика уровня зарплат по годам для выбранной профессии: {}'.format(vac_salary_years))
        print('Динамика количества вакансий по годам для выбранной профессии: {}'.format(vac_count_years))
        print('Уровень зарплат по городам (в порядке убывания): {}'.format(InputParam.first_corr(res_sort_area_salary)))
        print(
            'Доля вакансий по городам (в порядке убывания): {}'.format(InputParam.first_corr(res_sort_fract_vac_area)))

        return salary_years, vac_salary_years, vacs_years, vac_count_years, InputParam.first_corr(res_sort_area_salary), InputParam.first_corr(res_sort_fract_vac_area)


class Report:
    @staticmethod
    def weight(ws):
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

    @staticmethod
    def border(ws):
        thin = Side(border_style='thin', color="000000")
        for i, column in enumerate(ws.columns):
            for cell in column:
                if cell.value != '':
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    @staticmethod
    def get_head(ws, columns):
        for i, column in enumerate(columns):
            ws.cell(row=1, column=(1+i), value=column).font = Font(bold=True)

    @staticmethod
    def get_column_year(ws, info):
        salary_years = info[0]
        vac_salary_years = info[1]
        vacs_years = info[2]
        vac_count_years = info[3]
        for year, value in salary_years.items():
            ws.append([year, value, vac_salary_years[year], vacs_years[year], vac_count_years[year]])

    @staticmethod
    def get_column_area(ws, info):
        res_sort_area_salary = info[4]
        res_sort_fract_vac_area = info[5]
        for i, area in enumerate(res_sort_area_salary.keys()):
            ws.cell(row=2 + i, column=1).value = area
            ws.cell(row=2 + i, column=3).value = ''
        for i, area in enumerate(res_sort_fract_vac_area.keys()):
            ws.cell(row=2 + i, column=4).value = area
        for i, salary in enumerate(res_sort_area_salary.values()):
            ws.cell(row=2 + i, column=2).value = salary
        for i, fraction in enumerate(res_sort_fract_vac_area.values()):
            ws.cell(row=2 + i, column=5, value=fraction).number_format = BUILTIN_FORMATS[10]

    @staticmethod
    def report_excel(info):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Статистика по годам'
        Report.get_head(ws1, dicHeadFirst)
        Report.get_column_year(ws1, info)
        Report.weight(ws1)
        Report.border(ws1)

        ws2 = wb.create_sheet('Статистика по городам')
        Report.get_head(ws2, dicHeadSecond)
        Report.get_column_area(ws2, info)
        Report.weight(ws2)
        Report.border(ws2)
        wb.save('report.xlsx')


inp = InputParam()
data = DataSet.prepare(inp.params[0])
info = InputParam.print(data, inp.params[1])
dicHeadFirst = ['Год', 'Средняя зарплата', f'Средняя зарплата - {inp.params[1]}', 'Количество вакансий',
            f'Количество вакансий - {inp.params[1]}']
Report.report_excel(info)
